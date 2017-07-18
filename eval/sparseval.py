import getopt
import os
import sys

import subprocess

import openpyxl


def main(argv):
    bench_dir = ''
    output_file = ''
    timeout = 0
    try:
        opts, args = getopt.getopt(argv, "ht:d:o:", ["timeout=", "dir=", "output="])
    except getopt.GetoptError:
        print('invalid argument')
        print('sparseval.py -d [bench_dir] -o [output_file]')
        sys.exit(1)
    for opt, arg in opts:
        if opt == '-h':
            print('sparseval.py help:')
            print('sparseval.py -d [bench_dir] -o [output_file]')
            sys.exit(0)
        elif opt in ('-t', '--timeout'):
            timeout = int(arg)
        elif opt in ('-d', '--dir'):
            bench_dir = arg
            if not os.path.isdir(bench_dir):
                print('invalid benchmark dir')
                sys.exit(1)
        elif opt in ('-o', '--output'):
            output_file = arg
            ext_name = os.path.splitext(output_file)[1]
            if not (ext_name in ('.xls', '.xlsx')):
                print('invalid output file')
                sys.exit(1)
    evaluate(timeout, bench_dir, output_file)
    print('evaluation completed!')


def evaluate(timeout, bench_dir, output_file):
    args = ['python3', '/home/tsmart/dev/chengxi/PZ3/eval/sparsecounter.py']
    for root, dirs, files in os.walk(bench_dir):
        smt_files = [os.path.join(root, f) for f in files if f.endswith('.smt2')]
        for smt_file in smt_files:
            print(smt_file)
            args.append(smt_file)
            sparse = '*'
            factor = '*'
            try:
                result = subprocess.run(args, stdout=subprocess.PIPE, timeout=timeout)
                if result.returncode == 0:
                    sparse, factor = parse_result(result.stdout)
            except subprocess.TimeoutExpired:
                sparse = '*'
                factor = '*'
            export_result(smt_file, sparse, factor, output_file)
            args.pop()


def parse_result(output_str):
    lines = output_str.decode('utf-8').split('\n')
    one_line = lines[0]
    values = one_line.split(',')
    if len(values) == 2:
        sparse = values[0]
        factor = values[1]
    else:
        sparse = '*'
        factor = '*'
    return sparse, factor


def export_result(smt_file, sparse, factor, output_file):
    if not os.path.isfile(output_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        row_pointer = 1
    else:
        wb = openpyxl.load_workbook(filename=output_file)
        ws = wb.active
        row_pointer = ws.max_row + 1
    ws.cell(row=row_pointer, column=1).value = smt_file
    ws.cell(row=row_pointer, column=2).value = sparse
    ws.cell(row=row_pointer, column=3).value = factor
    wb.save(filename=output_file)


if __name__ == "__main__":
    main(sys.argv[1:])
