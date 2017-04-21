import os
import re
import sys
import getopt
import subprocess
import openpyxl


def main(argv):
    tool = ''
    bench_dir = ''
    core = 1
    timeout = 0
    export_stat = ''
    try:
        opts, args = getopt.getopt(argv, "hs:d:c:t:o:", ["solver=", "dir=", "core=", "timeout=", "output="])
    except getopt.GetoptError:
        print('invalid argument')
        print('scala.py -s <solver path> -d <benchmark dir> -c <maximum cores> -t <timeout in seconds> -o <export file>')
        sys.exit(1)
    for opt, arg in opts:
        if opt == '-h':
            print("script help:")
            print('scala.py -s <solver path> -d <benchmark dir> -c <maximum cores> -t <timeout in seconds> -o <export file>')
            sys.exit(0)
        elif opt in ("-s", "--solver"):
            tool = arg
        elif opt in ("-d", "--dir"):
            bench_dir = arg
        elif opt in ("-c", "--core"):
            core = int(arg)
        elif opt in ("-t", "--timeout"):
            timeout = int(arg)
        elif opt in ("-o", "--output"):
            export_stat = arg
            # the specified output statistics file must be .xls or .xlsx
            ext_name = os.path.splitext(export_stat)[1]
            if not (ext_name in (".xls", ".xlsx")):
                print('invalid output statistics file')
                sys.exit(1)
    if (not tool) or (not bench_dir) or (not export_stat):
        print('insufficient information for automatic evaluation')
        print('scala.py -s <solver path> -d <benchmark dir> -c <maximum cores> -t <timeout in seconds> -o <export file>')
        sys.exit(1)
    evaluate(tool, bench_dir, core, timeout, export_stat)
    print('evaluation completed!')


def evaluate(tool, bench_dir, core, timeout, export_stat):
    raw_result = []
    args = [tool]
    timeout_value = timeout if timeout > 0 else None
    for root, dirs, files in os.walk(bench_dir):
        smt_files = [os.path.join(root, f) for f in files if f.endswith(".smt2")]
        for smt_file in smt_files:
            args.append(smt_file)
            args.append(str(core))
            try:
                result = subprocess.run(args, stdout=subprocess.PIPE, timeout=timeout_value)
                if result.returncode == 0:
                    duration = load_duration(result.stdout)
                    raw_result.append((smt_file, duration))
            except subprocess.TimeoutExpired:
                # if we have triggered timeout exception, the specified timeout must be greater than 0
                # timeout value is in seconds, but we should store duration in milliseconds
                raw_result.append((smt_file, timeout_value * 1000))
            # after finished a case, we write the partial result to the file in case that the machine stops working
            # unexpectedly
            export_result(raw_result, export_stat)
            raw_result.clear()
            print(smt_file)
            args.pop()
            args.pop()
    return raw_result

time_pattern = re.compile(r"(\w+): (\d+)\D+")


def load_duration(output_str):
    decompose_time = ''
    subsolve_time = ''
    conciliate_time = ''
    lines = output_str.decode('utf-8').split('\n')
    for line in lines:
        match = re.match(time_pattern, line)
        if match:
            time_name = match.group(1)
            time_metric = match.group(2)
            if time_name == 'DECOMPOSITION':
                decompose_time = int(time_metric)
            elif time_name == 'SUBSOLVE':
                subsolve_time = int(time_metric)
            elif time_name == 'CONCILIATION':
                conciliate_time = int(time_metric)
    return decompose_time, subsolve_time, conciliate_time

case_name_column = 1
decompose_column = 2
subsolve_column = 3
conciliate_column = 4


def export_result(raw_result, export_stat):
    if not os.path.isfile(export_stat):
        wb = openpyxl.Workbook()
        ws = wb.active
        row_pointer = 1
    else:
        wb = openpyxl.load_workbook(filename=export_stat)
        ws = wb.active
        row_pointer = ws.max_row + 1
    for result in raw_result:
        case_name, solve_time = result
        decompose_time, subsolve_time, conciliate_time = solve_time
        ws.cell(row=row_pointer, column=case_name_column).value = case_name
        ws.cell(row=row_pointer, column=decompose_column).value = decompose_time
        ws.cell(row=row_pointer, column=subsolve_column).value = subsolve_time
        ws.cell(row=row_pointer, column=conciliate_column).value = conciliate_time
        row_pointer += 1
    wb.save(filename=export_stat)

if __name__ == "__main__":
    main(sys.argv[1:])
