import getopt
import os
import re
import sys

import subprocess

import openpyxl


def main(argv):
    tool = ''
    bench_dir = ''
    core = 2
    timeout = 0
    export_stat = ''
    try:
        opts, args = getopt.getopt(argv, "hs:d:c:t:o:", ["solver=", "dir=", "core=", "timeout=", "output="])
    except getopt.GetoptError:
        print('invalid argument')
        print('finegrained.py -s [solver] -d [benchmark dir] -c [max cores] -t [timeout] -o [export result]')
        sys.exit(1)
    for opt, arg in opts:
        if opt == '-h':
            print("script help:")
            print('finegrained.py -s [solver] -d [benchmark dir] -c [max cores] -t [timeout] -o [export result]')
            sys.exit(0)
        elif opt in ("-s", "--solver"):
            tool = arg
        elif opt in ("-d", "--dir"):
            bench_dir = arg
        elif opt in ("-c", "--core"):
            core = int(arg)
        elif opt in ("-t", "--timeout"):
            timeout = int(arg)
        elif opt in ("-o", "--output"):
            export_stat = arg
            ext_name = os.path.splitext(export_stat)[1]
            if not (ext_name in (".xls", ".xlsx")):
                print('invalid output file')
                sys.exit(1)
    if (not tool) or (not bench_dir) or (not export_stat):
        print('insufficient argument')
        print('finegrained.py -s [solver] -d [benchmark dir] -c [max cores] -t [timeout] -o [export result]')
        sys.exit(1)
    evaluate(tool, bench_dir, core, timeout, export_stat)
    print('evaluation completed!')


def evaluate(tool, bench_dir, core, timeout, export_stat):
    raw_result = []
    args = [tool]
    timeout_value = timeout if timeout > 0 else None
    for root, dirs, files in os.walk(bench_dir):
        smt_files = [os.path.join(root, f) for f in files if f.endswith(".smt2")]
        for smt_file in smt_files:
            args.append(smt_file)
            args.append(str(core))
            try:
                result = subprocess.run(args, stdout=subprocess.PIPE, timeout=timeout_value)
                if result.returncode == 0:
                    duration = load_duration(result.stdout)
                    raw_result.append((smt_file, duration))
            except subprocess.TimeoutExpired:
                # in this case, we discard partial results if any
                raw_result.append((smt_file, ('*', '*', '*', '*', '*', '*')))
            export_result(raw_result, export_stat)
            raw_result.clear()
            print(smt_file)
            args.pop()
            args.pop()

time_pattern = re.compile(r"(\w+): (\d+)")


def load_duration(output_str):
    decomp_time = '*'
    solve_time = '*'
    interp_time = '*'
    formulate_time = '*'
    ssr_time = '*'
    gs_time = '*'
    lines = output_str.decode('utf-8').split('\n')
    for line in lines:
        match = re.match(time_pattern, line)
        if match:
            time_name = match.group(1)
            time_metric = match.group(2)
            if time_name == 'DECOMP':
                decomp_time = int(time_metric)
            elif time_name == 'SOLVE':
                solve_time = int(time_metric)
            elif time_name == 'INTERP':
                interp_time = int(time_metric)
            elif time_name == 'SSR':
                ssr_time = int(time_metric)
            elif time_name == 'FORM':
                formulate_time = int(time_metric)
            elif time_name == 'GENSOLVE':
                gs_time = int(time_metric)
    return decomp_time, solve_time, interp_time, formulate_time, ssr_time, gs_time

case_name_column = 1
decomp_column = 2
solve_column = 3
interp_column = 4
formulate_column = 5
ssr_column = 6
gs_column = 7


def export_result(raw_result, export_stat):
    if not os.path.isfile(export_stat):
        wb = openpyxl.Workbook()
        ws = wb.active
        row_pointer = 1
    else:
        wb = openpyxl.load_workbook(filename=export_stat)
        ws = wb.active
        row_pointer = ws.max_row + 1
    for result in raw_result:
        case_name, time_result = result
        decomp_time, solve_time, interp_time, formulate_time, ssr_time, gs_time = time_result
        ws.cell(row=row_pointer, column=case_name_column).value = case_name
        ws.cell(row=row_pointer, column=decomp_column).value = decomp_time
        ws.cell(row=row_pointer, column=solve_column).value = solve_time
        ws.cell(row=row_pointer, column=interp_column).value = interp_time
        ws.cell(row=row_pointer, column=formulate_column).value = formulate_time
        ws.cell(row=row_pointer, column=ssr_column).value = ssr_time
        ws.cell(row=row_pointer, column=gs_column).value = gs_time
        row_pointer += 1
    wb.save(filename=export_stat)


if __name__ == "__main__":
    main(sys.argv[1:])
