import os
import re
import sys
import getopt
import subprocess
import openpyxl


def main(argv):
    tool = ''
    bench_dir = ''
    max_core = 1
    timeout = 0
    export_stat = ''
    try:
        # under this mode, if core=N, we run the solver on 2..N cores with timeout 2*TO..N*TO
        opts, args = getopt.getopt(argv, "hs:d:c:t:o:", ["solver=", "dir=", "core=", "timeout=", "output="])
    except getopt.GetoptError:
        print('invalid argument')
        print('scala.py -s <solver path> -d <benchmark dir> -c <maximum cores> -t <timeout in seconds> -o <export file>')
        sys.exit(1)
    for opt, arg in opts:
        if opt == '-h':
            print("script help:")
            print('scala.py -s <solver path> -d <benchmark dir> -c <maximum cores> -t <timeout in seconds> -o <export file>')
            sys.exit(0)
        elif opt in ("-s", "--solver"):
            tool = arg
        elif opt in ("-d", "--dir"):
            bench_dir = arg
        elif opt in ("-c", "--core"):
            max_core = int(arg)
        elif opt in ("-t", "--timeout"):
            timeout = int(arg)
        elif opt in ("-o", "--output"):
            export_stat = arg
            # the specified output statistics file must be .xls or .xlsx
            ext_name = os.path.splitext(export_stat)[1]
            if not (ext_name in (".xls", ".xlsx")):
                print('invalid output statistics file')
                sys.exit(1)
    if (not tool) or (not bench_dir) or (not export_stat):
        print('insufficient information for automatic evaluation')
        print('scala.py -s <solver path> -d <benchmark dir> -c <maximum cores> -t <timeout in seconds> -o <export file>')
        sys.exit(1)
    evaluate(tool, bench_dir, max_core, timeout, export_stat)
    print('evaluation completed!')


def evaluate(tool, bench_dir, max_core, timeout, export_stat):
    raw_result = []
    args = ['taskset','-c','0', tool]
    for root, dirs, files in os.walk(bench_dir):
        smt_files = [os.path.join(root, f) for f in files if f.endswith(".smt2")]
        for smt_file in smt_files:
            args.append(smt_file)
            for num_core in range(2, max_core + 1):
                args.append(str(num_core))
                timeout_value = timeout * num_core if timeout > 0 else None
                try:
                    result = subprocess.run(args, stdout=subprocess.PIPE, timeout=timeout_value)
                    if result.returncode == 0:
                        duration = load_duration(result.stdout)
                        raw_result.append((smt_file, num_core, duration))
                except subprocess.TimeoutExpired:
                    # if we have triggered timeout exception, the specified timeout must be greater than 0
                    raw_result.append((smt_file, num_core, timeout_value))
                args.pop()
            # after finished a case, we write the partial result to the file in case that the machine stops working
            # unexpectedly
            export_result(raw_result, export_stat)
            raw_result.clear()
            args.pop()
    return raw_result

time_pattern = re.compile(r"TOTAL: (\d+)\D+")


def load_duration(output_str):
    total_time = ''
    lines = output_str.decode('utf-8').split('\n')
    for line in lines:
        match = re.match(time_pattern, line)
        if match:
            total_time = int(match.group(1))
            break
    return total_time

case_name_column = 1
core_num_column = 2
solve_time_column = 3


def export_result(raw_result, export_stat):
    if not os.path.isfile(export_stat):
        wb = openpyxl.Workbook()
        ws = wb.active
        row_pointer = 1
    else:
        wb = openpyxl.load_workbook(filename=export_stat)
        ws = wb.active
        row_pointer = ws.max_row + 1
    for result in raw_result:
        case_name, core_num, solve_time = result
        ws.cell(row=row_pointer, column=case_name_column).value = case_name
        ws.cell(row=row_pointer, column=core_num_column).value = core_num
        ws.cell(row=row_pointer, column=solve_time_column).value = solve_time
        row_pointer += 1
    wb.save(filename=export_stat)

if __name__ == "__main__":
    main(sys.argv[1:])
