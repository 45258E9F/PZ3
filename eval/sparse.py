import getopt
import os
import sys
import openpyxl
import time

import subprocess


def main(argv):
    tool = ''
    bench_dir = ''
    num_core = 1
    timeout = 0
    export_file = ''
    try:
        opts, args = getopt.getopt(argv, "hs:d:c:t:o:", ["solver=", "dir=", "core=", "timeout=", "output="])
    except getopt.GetoptError:
        print('invalid argument')
        print('sparse.py -s [solver] -d [benchmark_dir] -c [#core] -t [timeout] -o [output_file]')
        sys.exit(1)
    for opt, arg in opts:
        if opt == '-h':
            print('sparse.py help:')
            print('sparse.py -s [solver] -d [benchmark_dir] -c [#core] -t [timeout] -o [output_file]')
            sys.exit(0)
        elif opt in ('-s', '--solver'):
            tool = arg
        elif opt in ('-d', '--dir'):
            bench_dir = arg
        elif opt in ('-c', '--core'):
            num_core = int(arg)
        elif opt in ('-t', '--timeout'):
            timeout = int(arg)
        elif opt in ('-o', '--output'):
            export_file = arg
    if (not tool) or (not bench_dir) or (not export_file):
        print('invalid argument')
        print('sparse.py -s [solver] -d [benchmark_dir] -c [#core] -t [timeout] -o [output_file]')
        sys.exit(1)
    if not os.path.isdir(bench_dir):
        print('invalid benchmark directory')
        sys.exit(1)
    if not os.path.isfile(tool):
        print('invalid solver')
        sys.exit(1)
    evaluate(tool, bench_dir, num_core, timeout, export_file)
    print('evaluation completed')


def evaluate(tool, bench_dir, num_core, timeout, export_file):
    raw_result = []
    args = [tool]
    timeout_value = timeout if timeout > 0 else None
    for root, dirs, files in os.walk(bench_dir):
        xlsx_files = [os.path.join(root, f) for f in files if f == 'meta.xlsx']
        if len(xlsx_files) == 1:
            # then we have found the meta file, including case name and its sparseness info
            xlsx_file = xlsx_files[0]
            wb = openpyxl.load_workbook(filename=xlsx_file)
            ws = wb.active
            for row in range(1, ws.max_row + 1):
                case = ws.cell(row=row, column=1).value
                sparse = ws.cell(row=row, column=2).value
                const_fac = ws.cell(row=row, column=3).value
                # solve the case by sequential/parallel solver
                case_file = os.path.join(root, case)
                args.append(case_file)
                duration_list = []
                for core in [1, num_core]:
                    args.append(str(core))
                    try:
                        start_time = time.time()
                        subprocess.run(args, stdout=subprocess.PIPE, timeout=timeout_value)
                        duration = int(round((time.time() - start_time) * 1000.0))
                        duration_list.append(duration)
                    except subprocess.TimeoutExpired:
                        duration_list.append('*')
                    args.pop()
                raw_result.append((case_file, sparse, const_fac, duration_list))
                print(case_file)
                args.pop()
            # export the statistics for each xlsx file
            export_result(raw_result, export_file)
            raw_result.clear()


def export_result(raw_result, export_file):
    if not os.path.isfile(export_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        row_pointer = 1
    else:
        wb = openpyxl.load_workbook(filename=export_file)
        ws = wb.active
        row_pointer = ws.max_row + 1
    for result in raw_result:
        case_file, sparse, const_fac, duration_list = result
        ws.cell(row=row_pointer, column=1).value = case_file
        ws.cell(row=row_pointer, column=2).value = sparse
        ws.cell(row=row_pointer, column=3).value = const_fac
        ws.cell(row=row_pointer, column=4).value = duration_list[0]
        ws.cell(row=row_pointer, column=5).value = duration_list[1]
        row_pointer += 1
    wb.save(filename=export_file)


if __name__ == "__main__":
    main(sys.argv[1:])
