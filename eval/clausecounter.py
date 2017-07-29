import getopt
import os
import sys

import openpyxl

sys.path.append('/home/tsmart/dev/chengxi/z3/build/python/')
import z3


def main(argv):
    bench_dir = ''
    output_file = ''
    try:
        opts, args = getopt.getopt(argv, "hd:o:", ["dir=", "output="])
    except getopt.GetoptError:
        print('invalid argument')
        sys.exit(1)
    for opt, arg in opts:
        if opt == '-h':
            print('clausecounter.py -d [bench_dir] -o [output_file]')
            sys.exit(0)
        elif opt in ('-d', '--dir'):
            bench_dir = arg
            if not os.path.isdir(bench_dir):
                print('invalid benchmark dir')
                sys.exit(1)
        elif opt in ('-o', '--output'):
            output_file = arg
            ext_name = os.path.splitext(output_file)[1]
            if not (ext_name in ('.xls', '.xlsx')):
                print('invalid output file')
                sys.exit(1)
    evaluate(bench_dir, output_file)
    print('evaluation completed!')


def evaluate(bench_dir, output_file):
    raw_result = []
    for root, dirs, files in os.walk(bench_dir):
        smt_files = [os.path.join(root, f) for f in files if f.endswith('.smt2')]
        for smt_file in smt_files:
            print(smt_file)
            num_clause = count(smt_file)
            raw_result.append((smt_file, num_clause))
    export_result(raw_result, output_file)


def count(smt_file):
    formula = z3.parse_smt2_file(smt_file)
    tactic_simp = z3.With(z3.Tactic('simplify'), 'elim_and', True)
    tactic_total = z3.Then(tactic_simp, z3.Tactic('elim-term-ite'))
    tactic_total = z3.Then(tactic_total, z3.Tactic('tseitin-cnf'))
    goals = tactic_total(formula)
    if len(goals) == 1:
        goal = goals[0]
        return len(goal)
    else:
        return -1


def export_result(raw_result, output_file):
    if not os.path.isfile(output_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        row_pointer = 1
    else:
        wb = openpyxl.load_workbook(filename=output_file)
        ws = wb.active
        row_pointer = ws.max_row + 1
    for result in raw_result:
        file_name, num_clause = result
        ws.cell(row=row_pointer, column=1).value = file_name
        ws.cell(row=row_pointer, column=2).value = num_clause
        row_pointer += 1
    wb.save(filename=output_file)


if __name__ == '__main__':
    main(sys.argv[1:])
