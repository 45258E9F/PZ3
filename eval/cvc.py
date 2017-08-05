import getopt
import os
import sys
import subprocess
import time

import openpyxl


def main(argv):
    seq_tool = ''
    par_tool = ''
    num_core = 4
    bench_dir = ''
    timeout = 0
    export_file = ''
    try:
        opts, args = getopt.getopt(argv, "hs:p:c:d:t:o:", ["seq=", "par=", "core=", "dir=" "timeout=", "output="])
    except getopt.GetoptError:
        print('invalid argument')
        print('cvc.py -s [seq solver] -p [par solver] -c [#core] -d [bench dir] -t [timeout] -o [output file]')
        sys.exit(1)
    for opt, arg in opts:
        if opt == '-h':
            print('cvc.py help:')
            print('cvc.py -s [seq solver] -p [par solver] -c [#core] -d [bench dir] -t [timeout] -o [output file]')
            sys.exit(0)
        elif opt in ('-s', '--seq'):
            seq_tool = arg
            if not os.path.isfile(seq_tool):
                print('invalid sequential solver')
                sys.exit(1)
        elif opt in ('-p', '--par'):
            par_tool = arg
            if not os.path.isfile(par_tool):
                print('invalid parallel solver')
                sys.exit(1)
        elif opt in ('-c', '--core'):
            num_core = int(arg)
        elif opt in ('-d', '--dir'):
            bench_dir = arg
            if not os.path.isdir(bench_dir):
                print('invalid benchmark directory')
                sys.exit(1)
        elif opt in ('-t', '--timeout'):
            timeout = int(arg)
        elif opt in ('-o', '--output'):
            export_file = arg
            ext_name = os.path.splitext(export_file)[1]
            if not (ext_name in ('.xls','.xlsx')):
                print('invalid output file')
                sys.exit(1)
    if (not seq_tool) or (not par_tool) or (not bench_dir) or (not export_file):
        print('invalid argument')
        print('cvc.py -s [seq solver] -p [par solver] -c [#core] -d [bench dir] -t [timeout] -o [output file]')
        sys.exit(1)
    evaluate(seq_tool, par_tool, num_core, bench_dir, timeout, export_file)
    print('evaluation completed!')


def evaluate(seq_tool, par_tool, num_core, bench_dir, timeout, export_file):
    args = []
    candidate_config = ['--thread0=--random-freq=0.5 --condense-function-values',
                        '--thread1=--random-seed=30 --symmetry-breaker',
                        '--thread2=--restart-int-base=50 --uf-ss-eager-split',
                        '--thread3=--restart-int-inc=5.0 --uf-ss-simple-cliques']
    timeout_value = timeout if timeout > 0 else None
    for root, dirs, files in os.walk(bench_dir):
        smt_files = [os.path.join(root, f) for f in files if f.endswith('.smt2')]
        for smt_file in smt_files:
            raw_result = (smt_file,)
            # first, we evaluate the sequential solver
            args.append(seq_tool)
            args.append(smt_file)
            try:
                start_time = time.time()
                result = subprocess.run(args, stdout=subprocess.PIPE, timeout=timeout_value)
                if result.returncode == 0:
                    duration = int(round((time.time() - start_time) * 1000.0))
                    raw_result += (duration,)
            except subprocess.TimeoutExpired:
                raw_result += ('*',)
            # second, we evaluate the portfolio solver
            args.clear()
            args.append(par_tool)
            args.append("--threads=%d" % num_core)
            for i in range(0, min(len(candidate_config), num_core)):
                args.append(candidate_config[i])
            args.append('--filter-lemma-length=8')
            args.append(smt_file)
            try:
                start_time = time.time()
                result = subprocess.run(args, stdout = subprocess.PIPE, timeout=timeout_value)
                if result.returncode == 0:
                    duration = int(round((time.time() - start_time) * 1000.0))
                    raw_result += (duration,)
            except subprocess.TimeoutExpired:
                raw_result += ('*',)
            export_result(raw_result, export_file)
            print(smt_file)
            args.clear()


def export_result(raw_result, export_file):
    if not os.path.isfile(export_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        row_pointer = 1
    else:
        wb = openpyxl.load_workbook(filename=export_file)
        ws = wb.active
        row_pointer = ws.max_row + 1
    if len(raw_result) == 3:
        ws.cell(row=row_pointer, column=1).value = raw_result[0]
        ws.cell(row=row_pointer, column=2).value = raw_result[1]
        ws.cell(row=row_pointer, column=3).value = raw_result[2]
        wb.save(filename=export_file)


if __name__ == "__main__":
    main(sys.argv[1:])
