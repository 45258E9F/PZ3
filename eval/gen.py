import getopt
import os

import sys
import random

import math
from collections import defaultdict

import openpyxl


class Formula:
    """
    The class for representation of formula.
    A formula is a conjunction of clauses.
    """

    def __init__(self):
        self.clauses = []
        self.symbols = []
        self.sparse = 1.0
        self.const_factor = 0.0

    def set_sparse(self, sparse):
        self.sparse = sparse

    def set_const_factor(self, factor):
        self.const_factor = factor

    def add_clause(self, clause):
        self.clauses.append(clause)

    def set_symbols(self, symbols):
        self.symbols = symbols

    def print(self):
        shuffled_symbols = list(self.symbols)
        random.shuffle(shuffled_symbols)
        printed_clauses = []
        for clause in self.clauses:
            printed_clauses.append(clause.print(shuffled_symbols))
        return '(and %s)' % ' '.join(printed_clauses)


class Clause:
    """
    The class for storing clause, which consists of a list of (in)equalities.
    They are connected by logical-or.
    """

    def __init__(self):
        self.literals = []

    def add_literal(self, literal):
        self.literals.append(literal)

    def print(self, symbols):
        printed_literals = []
        for literal in self.literals:
            printed_literals.append(literal.print(symbols))
        return '(or %s)' % ' '.join(printed_literals)


class Literal:
    """
    The class for representing (in)equality literal.
    """

    def __init__(self, edge, is_eq):
        # edge is a binary tuple
        self.left = edge[0]
        self.right = edge[1]
        self.is_eq = is_eq

    def print(self, symbols):
        result = '(= x%d x%d)' % (symbols[self.left], symbols[self.right])
        if not self.is_eq:
            result = '(not %s)' % result
        return result


class Graph:
    """
    A class for representing graphs
    """
    def __init__(self):
        self._graph = defaultdict(set)
        self._counter = 0
        self._threshold = 0

    def add_edge(self, l, r):
        if self._counter >= self._threshold:
            return
        if r not in self._graph[l]:
            self._counter += 1
            self._graph[l].add(r)
            self._graph[r].add(l)

    def set_threshold(self, threshold):
        self._threshold = threshold

    def is_saturated(self):
        return self._counter >= self._threshold

    def get_edges(self):
        edges = []
        for key in self._graph:
            for value in self._graph[key]:
                if key < value:
                    edges.append((key, value))
        return edges


def main(argv):
    num_clause = 0
    edge_size = 0
    const_factor = 0.0
    num_formula = 0
    output_dir = ''
    try:
        opts, args = getopt.getopt(argv, "hn:c:e:f:o:",
                                   ["num=", "clause=", "edge=", "factor=", "output="])
    except getopt.GetoptError:
        print('invalid argument')
        print(
            'gen.py -n [#problem] -c [#clause] -e [#(in)equality] -f [factor] -o [output dir]')
        sys.exit(1)
    for opt, arg in opts:
        if opt == '-h':
            print('script help:')
            print(
                'gen.py -n [#problem] -c [#clause] -e [#(in)equality] -f [factor] -o [output dir]')
            sys.exit(0)
        elif opt in ('-n', '--num'):
            num_formula = int(arg)
        elif opt in ('-c', '--clause'):
            num_clause = int(arg)
        elif opt in ('-e', '--edge'):
            edge_size = int(arg)
        elif opt in ('-f', '--factor'):
            const_factor = float(arg)
        elif opt in ('-o', '--output'):
            output_dir = arg
    if not output_dir:
        print('output directory not specified')
        sys.exit(1)
    if num_clause < 2:
        print('number of clause should be at least 3')
        sys.exit(1)
    if const_factor > 1.0 or const_factor < 0.0:
        print('factor should be between 0.0 and 1.0')
        sys.exit(1)
    if not os.path.isdir(output_dir):
        print('invalid output directory specified')
        sys.exit(1)
    gen(num_formula, num_clause, edge_size, const_factor, output_dir)
    print('gen completed')


def gen(num_formula, num_clause, edge_size, const_factor, output_dir):
    # create an excel file recording the association of test case and sparseness
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(0, num_formula):
        problem = gen_sub(num_clause, edge_size, const_factor)
        output_formula(problem, i, output_dir, ws)
    # output the workbook
    export_file = os.path.join(output_dir, 'meta.xlsx')
    wb.save(filename=export_file)


def gen_sub(num_clause, edge_size, const_factor):
    formula = Formula()
    num_visible_edge = random.randint(edge_size, num_clause * edge_size)
    sparse = (num_visible_edge - edge_size) / (num_clause * edge_size - edge_size)
    formula.set_sparse(sparse)
    # step 1: generate edges
    candidate_edges, const_num = gen_edges(num_visible_edge, const_factor)
    formula.set_const_factor(const_factor)
    formula.set_symbols(list(range(0, const_num)))
    # step 2: generate clauses to the formula
    # this problem is equivalent to random overlapping sets generation
    edge_flags = [True] * num_visible_edge
    num_covered_edge = 0
    num_rem_edge = num_clause * edge_size
    for i in range(0, num_clause):
        clause = Clause()
        # invariant: num_rem_edge >= num_visible_edge
        sample_size_covered = random.randint(max(0, edge_size - num_visible_edge),
                                             min(num_rem_edge - num_visible_edge, edge_size, num_covered_edge))
        sample_size_uncovered = edge_size - sample_size_covered
        if sample_size_covered > 0:
            covered = [i for i, e in enumerate(edge_flags) if not e]
            covered_hits = random.sample(covered, sample_size_covered)
            for covered_hit in covered_hits:
                clause.add_literal(Literal(candidate_edges[covered_hit], bool(random.getrandbits(1))))
        if sample_size_uncovered > 0:
            num_covered_edge += sample_size_uncovered
            num_visible_edge -= sample_size_uncovered
            uncovered = [i for i, e in enumerate(edge_flags) if e]
            uncovered_hits = random.sample(uncovered, sample_size_uncovered)
            for uncovered_hit in uncovered_hits:
                edge_flags[uncovered_hit] = False
                clause.add_literal(Literal(candidate_edges[uncovered_hit], bool(random.getrandbits(1))))
        num_rem_edge -= edge_size
        formula.add_clause(clause)
    return formula


def gen_edges(num_visible_edge, const_factor):
    const_min = math.ceil((1 + math.sqrt(1 + 8 * num_visible_edge)) / 2)
    const_max = 2 * num_visible_edge
    delta = round(const_factor * (const_max - const_min))
    const_num = const_min + delta
    # add basic edges first
    graph = Graph()
    graph.set_threshold(num_visible_edge)
    for i in range(0, const_num, 2):
        graph.add_edge((i + 1) % const_num, i)
    if not graph.is_saturated():
        # sample the remaining edges
        # sample `num_visible_edge` edges first
        total_edge_num = int((const_num - 1) * const_num / 2)
        candidates = random.sample(range(1, total_edge_num + 1), num_visible_edge)
        # decode the candidate number to the edge
        for candidate in candidates:
            # calculate the row and column of certain tuple
            row = math.ceil((math.sqrt(1 + 8 * candidate) - 1) / 2)
            column = int(candidate - (row - 1) * row / 2) - 1
            # add them to the graph
            graph.add_edge(row, column)
    return graph.get_edges(), const_num


def output_formula(problem, i, output_dir, ws):
    file_name = "problem_" + str(i) + ".smt2"
    # add workbook entry
    ws.cell(row=i + 1, column=1).value = file_name
    ws.cell(row=i + 1, column=2).value = problem.sparse
    ws.cell(row=i + 1, column=3).value = problem.const_factor
    file_path = os.path.join(output_dir, file_name)
    f = open(file_path, 'w')
    f.write('(declare-sort U 0)\n')
    # declare constants
    for j in problem.symbols:
        f.write('(declare-fun x%d () U)\n' % j)
    # assert the constraint
    f.write('(assert %s)\n' % problem.print())
    f.write('(check-sat)\n')
    f.write('(exit)\n')
    f.close()


if __name__ == "__main__":
    main(sys.argv[1:])
